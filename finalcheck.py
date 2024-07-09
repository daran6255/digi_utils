from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string



def highlight_cells(input_file,output_file):
    
    def capitalize_first_letter(cell_value):
        if cell_value:
            words = cell_value.split()
            capitalized_words = [word.capitalize() for word in words]
            return ' '.join(capitalized_words)
        return cell_value
    # Load the workbook
    wb = load_workbook(input_file)
    ws = wb.active
    column_index = 'A'
    values_to_highlight = [
        'AC', 'Autoclave', 'Baby Warmer', 'Battery', 'Battery + Inverter', 'Blood Mixer', 'Cbc Machine',
        'CCTV', 'Ceiling Fan', 'Cell Counter', 'Centrifuge Machine', 'CFL Bulb', 'CFL Tubelight', 
        'Computer,Desktop', 'Deep Freezer', 'Digital Clock', 'Digital Weighing Machine', 'Electric Kettle',
        'Exhaust Fan', 'Flood Light', 'Generator', 'Heater', 'Hematology Analyzer', 'Hot Air Oven',
        'Hub Cutter (Needle)', 'Ice Lined Refrigerator (ILR)', 'Incandescent Bulb', 'Incubator', 'Inverter',
        'Laptop', 'LED Bulb', 'LED Tube Light', 'Microscope', 'Mixer Machine', 'Mobile Charging Point',
        'Nebulizer Machine', 'Needle Cum - Syringe Burner/ Destroyer', 'Needle Cutter', 'Operation Theratre/ OT Light',
        'Oxygen Concentrator', 'Patient Monitor', 'Pedestal Fan', 'Printer', 'Refrigerator /Fridge',
        'Road/Street Light', 'Semi Autoanalyser', 'Solar inverter', 'Solar Panel', 'Spot Light', 'Sterilizer',
        'Suction Machine', 'T.V', 'Table Fan', 'Tablet Pc', 'Truenat', 'Vacuum Cleaner', 'Voltage Stabilizer',
        'Wall Fan', 'Water Pump', 'Water Purifer', 'Water/Air Cooler', 'X-Ray Mechine', 'X-Ray View Box',
        ]
    col_index = column_index_from_string(column_index)
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                words = cell_value.split()
                capitalized_words = [word[0].upper() + word[1:].lower() for word in words]
                ws.cell(row=row, column=col).value = ' '.join(capitalized_words)
                 
        # Highlight cells with non-matching values
    fill = PatternFill(start_color='FBD5AB', end_color='FBD5AB', fill_type='solid')
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col_index).value
        if cell_value not in values_to_highlight:
            ws.cell(row=row, column=col_index).fill = fill

    # Save the workbook
    wb.save(output_file)
    print(f"Cells in column '{column_index}' with non-matching values highlighted and saved to '{output_file}'")
    return output_file


# Highlight cells with non-matching values in the specified column