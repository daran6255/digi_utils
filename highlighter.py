import openpyxl
import re
import random
from openpyxl.utils import column_index_from_string

def load_workbooks(file1, file2):
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)
    return wb1, wb2

def select_sheets(wb1, wb2):
    sheet1 = wb1.active
    sheet2 = wb2.active
    return sheet1, sheet2

def clean_text(cell):
    if isinstance(cell.value, str):
        pattern1 = r"Im stopping the model because I couldnt find any brand name matching nan"
        pattern2 = r"Im stopping the model since I couldnt find any brand name matching nan"
        pattern3 = r"nan"
        pattern4 = r"Nan"
        cell.value = re.sub(pattern1, '', cell.value)
        cell.value = re.sub(pattern2, '', cell.value)
        cell.value = re.sub(pattern3, '', cell.value)
        cell.value = re.sub(pattern4, '', cell.value)

def clean_sheets(sheet1, sheet2):
    for ws in [sheet1, sheet2]:
        for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, sheet2.max_row), min_col=1, max_col=ws.max_column):
            for cell in row:
                clean_text(cell)

def compare_and_highlight(sheet1, sheet2, columns_to_compare):
    mismatch_color = 'ff6666'  # Red
    single_value_color = '00e6e6'  # Green
    space_highlight_color = 'e6e600'  # Yellow

    for row1, row2 in zip(sheet1.iter_rows(), sheet2.iter_rows()):
        non_null_cells = [cell for cell in row2 if cell.value is not None and str(cell.value).strip()!= '']
        if len(non_null_cells) == 1:
            non_null_cells[0].fill = openpyxl.styles.PatternFill(start_color=single_value_color, end_color=single_value_color, fill_type='solid')
        for cell1, cell2 in zip(row1, row2):
            column_index = cell1.column
            column_letter = openpyxl.utils.get_column_letter(column_index)
            if column_letter in columns_to_compare:
                if cell1.value is None or cell2.value is None:
                    continue
                if str(cell1.value)!= str(cell2.value):
                    cell2.fill = openpyxl.styles.PatternFill(start_color=mismatch_color, end_color=mismatch_color, fill_type='solid')
            if column_letter == 'D':  # Check for spaces in column D
                if cell2.value is not None:
                    if not re.search(r'\s', str(cell2.value).strip()):  # If value doesn't have a space, highlight
                        cell2.fill = openpyxl.styles.PatternFill(start_color=space_highlight_color, end_color=space_highlight_color, fill_type='solid')

def remove_decimal_points(column_letter, wb2):
    column_index1 = column_index_from_string(column_letter)
    ws = wb2.active 
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=column_index1)  # Get the cell object
        if cell is not None:  # Check if the cell exists
            cell_value = cell.value  # Get the cell value
            if isinstance(cell_value, (str, float)):  # Check if the cell value is a string or float
                cell.value = str(cell_value).replace('.', '')
                
                        
def highlight(file1,file2,output_file,columns_to_compare):
    
    columns_to_compare = [column.strip().upper() for column in columns_to_compare]
    wb1, wb2 = load_workbooks(file1, file2)
    remove_decimal_points('E',wb2)
    sheet1, sheet2 = select_sheets(wb1, wb2)
    clean_sheets(sheet1, sheet2)
    compare_and_highlight(sheet1, sheet2, columns_to_compare)
    wb2.save(output_file)
